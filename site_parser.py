from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
import time
import os
import sys
from pathlib import Path
from logger import setup_logger

# Создаем логгер для парсера
parser_logger = setup_logger('parser', 'parser.log')

class SiteParser:
    def __init__(self, excel_dir='excel_reports'):
        """Инициализация парсера"""
        self.excel_dir = excel_dir
        # Создаем директорию для Excel отчетов
        Path(excel_dir).mkdir(parents=True, exist_ok=True)
        parser_logger.info(f"SiteParser initialized, Excel directory: {excel_dir}")
        
    def _get_driver(self):
        """Создает и возвращает настроенный драйвер"""
        parser_logger.debug("Creating Chrome driver...")
        
        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        
        if sys.platform == "darwin":  # macOS
            options.add_argument("--disable-dev-shm-usage")
        
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options) # type: ignore
            driver.implicitly_wait(10)
            parser_logger.info("Chrome driver created successfully")
            return driver
        except Exception as e:
            parser_logger.error(f"Failed to create Chrome driver: {e}")
            raise
    
    def test_auth(self, user_login, user_password):
        """Проверяет авторизацию на сайте"""
        parser_logger.info(f"Testing auth for user: {user_login}")
        driver = self._get_driver()
        
        try:
            driver.get('http://82.116.39.87:82/overtime/logon.php')
            parser_logger.debug("Login page loaded")
            
            driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
            driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
            driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
            time.sleep(1)
            
            try:
                user_name = driver.find_element('xpath', '//html[1]/body[1]/div[1]/div[1]/table[2]/tbody[1]/tr[1]/td[1]').text
                parser_logger.info(f"Auth successful for {user_login}: {user_name}")
                return {
                    'success': True,
                    'message': f'Вы авторизованы на сайте как {user_name}',
                    'user_name': user_name
                }
            except NoSuchElementException:
                parser_logger.warning(f"Auth failed for {user_login}: wrong credentials")
                return {
                    'success': False,
                    'message': 'Не удалось авторизоваться на сайте\nПопробуйте зарегистрироваться сначала либо обратитесь к администратору'
                }
        except Exception as e:
            parser_logger.error(f"Auth error for {user_login}: {e}")
            return {
                'success': False,
                'message': f'Ошибка при авторизации: {str(e)}'
            }
        finally:
            driver.quit()
            parser_logger.debug("Driver closed")
    
    def get_total_km(self, user_login, user_password):
        """Возвращает общий пробег пользователя"""
        parser_logger.info(f"Getting total km for user: {user_login}")
        driver = self._get_driver()
        
        try:
            driver.get('http://82.116.39.87:82/overtime/logon.php')
            driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
            driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
            driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
            time.sleep(0.5)
            
            driver.get('http://82.116.39.87:82/overtime/mileage.php')
            time.sleep(1)
            
            kol_iter = driver.find_elements('xpath', '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr')
            total_km = 0
            
            for i in range(len(kol_iter) - 1):
                iter_num = i + 1
                path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[6]'
                number = driver.find_element('xpath', path).text
                if number.isdigit():
                    total_km += int(number)
            
            parser_logger.info(f"Total km for {user_login}: {total_km}, records: {len(kol_iter) - 1}")
            return {
                'success': True,
                'total_km': total_km,
                'records_count': len(kol_iter) - 1
            }
        except NoSuchElementException as e:
            parser_logger.error(f"NoSuchElementException for {user_login}: {e}")
            return {
                'success': False,
                'message': 'Ошибка при получении данных о пробеге'
            }
        except Exception as e:
            parser_logger.error(f"Error getting total km for {user_login}: {e}")
            return {
                'success': False,
                'message': f'Ошибка: {str(e)}'
            }
        finally:
            driver.quit()
    
    def get_mileage_report(self, user_login, user_password, chat_id):
        """Создает Excel отчет о пробеге"""
        parser_logger.info(f"Creating Excel report for user: {user_login}, chat_id: {chat_id}")
        driver = self._get_driver()
        
        try:
            driver.get('http://82.116.39.87:82/overtime/logon.php')
            driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
            driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
            driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
            time.sleep(0.5)
            
            driver.get('http://82.116.39.87:82/overtime/mileage.php')
            time.sleep(1)
            
            kol_iter = driver.find_elements('xpath', '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr')
            parser_logger.debug(f"Found {len(kol_iter)} records for {user_login}")
            
            # Создаем Excel файл
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            
            workbook = Workbook()
            sheet = workbook.active
            
            # Настройка ширины колонок
            sheet.column_dimensions["A"].width = 5 # type: ignore
            sheet.column_dimensions["B"].width = 14 # type: ignore
            sheet.column_dimensions["C"].width = 12 # type: ignore
            sheet.column_dimensions["D"].width = 20 # type: ignore
            sheet.column_dimensions["E"].width = 100 # type: ignore
            sheet.column_dimensions["F"].width = 8 # type: ignore
            sheet.column_dimensions["G"].width = 13 # type: ignore
            sheet.column_dimensions["H"].width = 100 # type: ignore
            
            # Стиль для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Заголовки
            headers = ["№", "ID", "Дата", "Населённый пункт", "Адрес", "КМ", "Куда", "Комментарий"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col) # type: ignore
                cell.value = header # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            total_km = 0
            records_processed = 0
            
            # Заполняем данные
            for i in range(len(kol_iter) - 1):
                row_num = i + 2
                iter_num = i + 1
                
                try:
                    # Номер
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[1]'
                    sheet.cell(row=row_num, column=1).value = driver.find_element('xpath', path).text # type: ignore
                    
                    # ID
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[2]'
                    sheet.cell(row=row_num, column=2).value = driver.find_element('xpath', path).text # type: ignore
                    
                    # Дата
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[3]'
                    sheet.cell(row=row_num, column=3).value = driver.find_element('xpath', path).text # type: ignore
                    sheet.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center") # type: ignore
                    
                    # Город
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[4]'
                    sheet.cell(row=row_num, column=4).value = driver.find_element('xpath', path).text # type: ignore
                    sheet.cell(row=row_num, column=4).alignment = Alignment(horizontal="center", vertical="center") # type: ignore
                    
                    # Адрес
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[5]'
                    sheet.cell(row=row_num, column=5).value = driver.find_element('xpath', path).text # type: ignore
                    
                    # КМ
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[6]'
                    km_value = driver.find_element('xpath', path).text
                    if km_value.isdigit():
                        sheet.cell(row=row_num, column=6).value = int(km_value) # type: ignore
                        total_km += int(km_value)
                    sheet.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="center") # type: ignore
                    
                    # Куда
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[7]'
                    sheet.cell(row=row_num, column=7).value = driver.find_element('xpath', path).text # type: ignore
                    
                    # Комментарий
                    path = f'//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[{iter_num}]/td[8]'
                    sheet.cell(row=row_num, column=8).value = driver.find_element('xpath', path).text # type: ignore
                    
                    records_processed += 1
                except Exception as e:
                    parser_logger.error(f"Error processing record {iter_num} for {user_login}: {e}")
                    continue
            
            # Добавляем итоговую строку
            last_row = records_processed + 2
            sheet.cell(row=last_row, column=5).value = "Общий пробег = " # type: ignore
            sheet.cell(row=last_row, column=5).alignment = Alignment(horizontal="right", vertical="center") # type: ignore
            sheet.cell(row=last_row, column=6).value = total_km # type: ignore
            sheet.cell(row=last_row, column=6).font = Font(bold=True) # type: ignore
            
            # Добавляем информацию о норме
            norm_row = last_row + 1
            expected_fuel = total_km / 10
            sheet.cell(row=norm_row, column=5).value = "При норме 10 л/100 км:" # type: ignore
            sheet.cell(row=norm_row, column=5).alignment = Alignment(horizontal="right", vertical="center") # type: ignore
            sheet.cell(row=norm_row, column=6).value = f"{expected_fuel:.1f} литров" # type: ignore
            
            # Сохраняем файл в директорию excel_reports
            timestamp = int(time.time())
            filename = f"mileage_{chat_id}_{timestamp}.xlsx"
            filepath = os.path.join(self.excel_dir, filename)
            workbook.save(filepath)
            
            parser_logger.info(f"Excel report created: {filename}, total_km={total_km}, records={records_processed}")
            return {
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'total_km': total_km,
                'records_count': records_processed
            }
            
        except NoSuchElementException as e:
            parser_logger.error(f"NoSuchElementException in report for {user_login}: {e}")
            return {
                'success': False,
                'message': 'Ошибка при получении данных о пробеге'
            }
        except Exception as e:
            parser_logger.error(f"Error creating report for {user_login}: {e}")
            return {
                'success': False,
                'message': f'Ошибка: {str(e)}'
            }
        finally:
            driver.quit()
            parser_logger.debug("Driver closed")